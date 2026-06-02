"""Export annotations to an .xlsx workbook.

Kept Qt-free so the rendering logic stays testable and the file_list_view
just passes in pre-formatted rows. Uses openpyxl which is bundled into the
PyInstaller build via requirements.txt.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


@dataclass(frozen=True)
class ExportRow:
    """One row in the exported xlsx, in display order."""

    video_number: str       # e.g. "00005" — the NNNNN suffix from the filename
    process_name: str       # the Process Name from the annotation, "" if unannotated
    duration_text: str      # formatted duration like "01:23" or "1:02:45", "" if unknown


def export_to_xlsx(rows: list[ExportRow], output_path: Path) -> None:
    """Write the rows to ``output_path`` as a single-sheet xlsx workbook.

    Overwrites any existing file at the path. Raises whatever openpyxl /
    filesystem error occurs; caller is expected to surface it to the user.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Annotations"

    headers = ["视频编号", "视频名称(工序名)", "视频拍摄时长"]
    ws.append(headers)
    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold
        cell.alignment = center

    for r in rows:
        ws.append([r.video_number, r.process_name, r.duration_text])

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 16

    # Right-align the duration column, center the video number column.
    # Force the video-number column to text format ("@"). Excel / WPS see
    # an unformatted cell with a value like "00002" and "helpfully" coerce
    # it to a number, dropping the leading zeros → "2". The text format
    # tells them to display the string verbatim.
    right = Alignment(horizontal="right")
    for row in range(2, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        cell_a.alignment = center
        cell_a.number_format = "@"
        ws.cell(row=row, column=3).alignment = right

    ws.freeze_panes = "A2"

    wb.save(str(output_path))
